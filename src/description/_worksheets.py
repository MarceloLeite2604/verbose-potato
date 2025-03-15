import json

from configuration import retrieve_output_file_path
from util.file import as_file_name
from workbook import load_workbook
from formulas import Parser
from openpyxl.worksheet.formula import ArrayFormula
from util.cell import split_cell_reference

_PARSER = Parser()


def _is_empty(cell):
    return not cell.value


def _is_input(cell):

    return cell.style \
        and cell.style == 'Editáveis'


def _retrieve_formula_inputs(formula):
    inputs = _PARSER.ast(formula)[1].compile().inputs

    if len(inputs) <= 0:
        return None

    formula_inputs = {}

    for key, _ in inputs.items():

        cell_reference = split_cell_reference(key)

        if cell_reference:
            input_type = 'ranges' if 'end' in cell_reference else 'cells'
        else:
            input_type = 'defined_names'

        if input_type not in formula_inputs:
            formula_inputs[input_type] = []

        formula_inputs[input_type].append(key)

    return formula_inputs


def _elaborate_cell_formula(cell):
    formula = {}

    if type(cell.value) == ArrayFormula:
        formula['definition'] = f'=ARRAYFORMULA({cell.value.text[1:]})'
    else:
        formula['definition'] = cell.value

    inputs = _retrieve_formula_inputs(formula['definition'])

    if inputs:
        formula['inputs'] = inputs

    return formula


def _elaborate_cell_data(cell):
    if _is_empty(cell) and not _is_input(cell):
        return None

    cell_data = {}

    match cell.data_type:
        case 'f':
            cell_data['formula'] = _elaborate_cell_formula(cell)

        case 's' | 'n' | 'b' | 'e' | 'd':
            if cell.value:
                cell_data['value'] = cell.value

    if _is_input(cell):
        cell_data['require_input'] = True

    return cell_data


def _elaborate_descriptions(definition):

    worksheet_title = definition['title']

    descriptions = {}

    for cell_range, cell_data in definition['cells'].items():
        description = f"Cell \"{worksheet_title}!{cell_range}\" "

        if 'formula' in cell_data:
            cell_type = 'formula'
            description += f"has formula \"{cell_data['formula']['definition']}\"."
        elif 'value' in cell_data:
            cell_type = 'value'
            description += f"has value \"{cell_data['value']}\"."
        elif 'require_input' in cell_data:
            cell_type = 'input'
            description += "requires input from user."

        metadata = {
            'cell_range': cell_range,
            'type': cell_type,
        }

        if cell_type == 'formula' \
                and 'inputs' in cell_data['formula']:

            inputs = cell_data['formula']['inputs']
            formula_inputs = []

            if 'ranges' in inputs:
                formula_inputs.extend(inputs['ranges'])

            if 'cells' in inputs:
                formula_inputs.extend(inputs['cells'])

            if 'defined_names' in inputs:
                formula_inputs.extend(inputs['defined_names'])

            metadata['formula-inputs'] = formula_inputs

        descriptions[cell_range] = {
            'text': description,
            'metadata': metadata
        }

    return descriptions


def _write_worksheet_definition(worksheet):
    print(f'Processing worksheet {worksheet.title}.')

    definition = {
        'title': worksheet.title
    }

    cells = {}

    for row_index, row in enumerate(worksheet.iter_rows()):

        if row_index != 0 and row_index % 10 == 0:
            print(f'Processing row {row_index}.')

        row_cells = {}

        for cell in row:
            cell_data = _elaborate_cell_data(cell)
            if cell_data:
                row_cells[cell.coordinate] = cell_data

        if len(row_cells.keys()) > 0:
            cells = cells | row_cells

    if len(cells.keys()) > 0:
        definition['cells'] = cells

    descriptions = _elaborate_descriptions(definition)

    file_name = f'{as_file_name(worksheet.title)}.json'

    output_path = retrieve_output_file_path(
        'definitions', 'worksheets', file_name)

    with open(output_path, 'w') as file:
        json.dump(descriptions, file, ensure_ascii=False,
                  indent=2, separators=(',', ': '))


def write_definitions():
    workbook = load_workbook()

    for index, worksheet in enumerate(workbook.worksheets):
        _write_worksheet_definition(worksheet)
