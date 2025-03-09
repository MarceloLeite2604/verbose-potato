import json
import os
import regex as re
from openpyxl import load_workbook
from dotenv import load_dotenv, find_dotenv

load_dotenv(find_dotenv(), override=True)

spreadsheet_location = os.environ['INPUT_FILE_PATH']
output_directory = 'output-files'
workbook_description_output_file = os.path.join(
    output_directory, 'workbook-description.txt')

if not os.path.exists(output_directory):
    os.makedirs(output_directory)


def is_empty(cell):
    return not cell.value


def is_input_cell(cell):
    return cell.fill.bgColor.rgb \
        and cell.fill.bgColor.rgb == 'FFFCD5B4'


def retrieve_defined_name_range(defined_name):
    range_regex = r'^\$?(?<start_col>[A-Z]+)\$?(?<start_row>\d+)(?::\$?(?<end_col>[A-Z]+)\$(?<end_row>\d+))?'

    for (worksheet_title, cell_range) in defined_name.destinations:

        if not cell_range:
            return None

        match = re.match(range_regex, cell_range)

        if not match:
            raise ValueError(
                f'Unknown range format \"{cell_range}\" for defined name \"{defined_name.name}\".')

        return {
            'worksheet_title': worksheet_title,
            'start': {
                'col': match['start_col'],
                'row': match['start_row']
            } if match['start_col']
            and match['start_row']
            else None,
            'end': {
                'col': match['end_col'],
                'row': match['end_row']
            } if match['end_col']
            and match['end_row']
            else None
        }


def retrieve_global_defined_names(workbook):

    defined_names = {}

    for defined_name, definition in workbook.defined_names.items():

        defined_name_range = retrieve_defined_name_range(definition)

        if defined_name_range:

            if defined_name_range['end']:
                if defined_name_range['start']['col'] == defined_name_range['end']['col'] \
                        or defined_name_range['start']['row'] == defined_name_range['end']['row']:
                    defined_name_type = 'list'
                else:
                    defined_name_type = 'table'
            else:
                defined_name_type = 'constant'

            defined_names[defined_name] = {
                'name': defined_name,
                'type': defined_name_type,
                'range': defined_name_range
            }

    return defined_names


def elaborate_workbook_data(file_path):
    workbook = load_workbook(file_path, data_only=False)

    workbook_data = {
        'worksheets': {}
    }

    # Retrieve workbook named ranges
    global_defined_names = retrieve_global_defined_names(workbook)

    workbook_data['defined_names'] = global_defined_names

    selected_worksheet_index = 4

    worksheet = workbook.worksheets[selected_worksheet_index]

    sheet_data = {
        'name': worksheet.title
    }

    cells = []

    for row in worksheet.iter_rows():

        for cell in row:

            if is_empty(cell) and not is_input_cell(cell):
                continue

            cell_data = {
                "coordinate": cell.coordinate
            }

            match cell.data_type:
                case 'f':
                    cell_data['formula'] = cell.value
                case 's' | 'n' | 'b' | 'e' | 'd':
                    if cell.value:
                        cell_data['value'] = cell.value

            # Print cell background color rgb value if defined
            if is_input_cell(cell):
                cell_data['require_input'] = True

            cells.append(cell_data)

    if len(cells) > 0:
        sheet_data['cells'] = cells

    workbook_data['worksheets'][selected_worksheet_index] = sheet_data

    return workbook_data


def elaborate_global_defined_names_description(defined_names):

    with open(workbook_description_output_file, 'w') as file:

        for name, data in defined_names.items():

            defined_name_type = data['type']

            description = f'The workbook has a {defined_name_type} named \"{name}\". '

            defined_name_range = data['range']

            match defined_name_type:
                case 'constant':
                    coordinate = f"'{defined_name_range['worksheet_title']}'!${defined_name_range['start']['col']}${defined_name_range['start']['row']}"
                    description += f"Its value is based on cell \"{coordinate}\"."
                case 'list' | 'table':
                    coordinate = f"'{defined_name_range['worksheet_title']}'!${defined_name_range['start']['col']}${defined_name_range['start']['row']}:${defined_name_range['end']['col']}${defined_name_range['end']['row']}"
                    description += f"Its value is based on range {coordinate}."

            file.write(description + '\n')


def elaborate_cells_description(worksheets_data):

    with open(workbook_description_output_file, 'a') as file:

        for worksheet_index, worksheet_data in worksheets_data.items():
            worksheet_name = worksheet_data['name']

            for cell in worksheet_data['cells']:

                cell_coordinates = f"'{worksheet_name}'!{cell['coordinate']}"

                if 'value' in cell:
                    cell_description = f"contains the value \"{cell['value']}\""

                if 'formula' in cell:
                    cell_description = f"contains the formula \"{cell['formula']}\""

                if 'require_output' in cell:
                    cell_description = 'requires an input from the user'

                cell_information = f'Cell {cell_coordinates} {cell_description}.'

                cell_information = cell_information.replace('\n', '\\n')\
                    .replace('\t', '\\t')

                file.write(cell_information + '\n')


def elaborate_workbook_description(workbook_data):

    elaborate_global_defined_names_description(workbook_data['defined_names'])

    elaborate_cells_description(workbook_data['worksheets'])


workbook_data = elaborate_workbook_data(spreadsheet_location)


output_file = os.path.join(output_directory, 'workbook-data.json')

with open(output_file, 'w') as file:
    json.dump(workbook_data, file, ensure_ascii=False,
              indent=2, separators=(',', ': '))

elaborate_workbook_description(workbook_data)
