import json
import os
import regex as re
from openpyxl import load_workbook
from configuration import load_env, OUTPUT_DIRECTORY
from formulas import Parser
import util.cell as cell_util

load_env()

COORDINATE_REGEX = r"^(\$?'?(?<workbook_name>[A-zÀ-ú ]+)'?[!\.])?\$?(?<start_col>[A-Z]+)\$?(?<start_row>\d+)(?::\$?(?<end_col>[A-Z]+)\$?(?<end_row>\d+))?$"

input_file_path = os.environ['INPUT_FILE_PATH']


DESCRIPTION_DIRECTORY = os.path.join(OUTPUT_DIRECTORY, 'description')

# global_defined_names_output_file = os.path.join(
#     DESCRIPTION_DIRECTORY, 'global-defined-names.json')

# workbook_description_output_file = os.path.join(
#     output_directory, 'workbook-description.json')


def elaborate_workbook_data(file_path):
    workbook = load_workbook(file_path, data_only=False)

    parser = Parser()

    workbook_data = {
        'worksheets': {}
    }

    write_global_defined_names_file(workbook)

    for selected_worksheet_index, worksheet in enumerate(workbook.worksheets):

        print(f'Processing worksheet {worksheet.title}.')

        sheet_data = {
            'name': worksheet.title
        }

        cells = []

        index = 0
        for row in worksheet.iter_rows():

            if index != 0 and index % 10 == 0:
                print(f'Processing row {index}.')

            index += 1

            for cell in row:

                if is_empty(cell) and not is_input_cell(cell):
                    continue

                cell_data = {
                    "coordinate": cell.coordinate
                }

                match cell.data_type:
                    case 'f':
                        formula = cell_data['formula'] = {}
                        if type(cell.value) == ArrayFormula:
                            formula['definition'] = f'=ARRAYFORMULA({cell.value.text[1:]})'
                        else:
                            formula['definition'] = cell.value

                        function = parser.ast(formula['definition'])[
                            1].compile()

                        if len(function.inputs) > 0:
                            inputs = formula['inputs'] = {}

                        for key, _ in function.inputs.items():

                            match = re.match(coordinate_regex, key)

                            if match:
                                if match['end_row'] \
                                        and match['end_col']:
                                    input_type = 'ranges'
                                else:
                                    input_type = 'cells'
                            else:
                                input_type = 'defined_names'

                            if input_type not in inputs:
                                inputs[input_type] = []

                            inputs[input_type].append(key)

                    case 's' | 'n' | 'b' | 'e' | 'd':
                        if cell.value:
                            cell_data['value'] = cell.value

                # Print cell background color rgb value if defined
                if is_input_cell(cell):
                    cell_data['require_input'] = True

                cells.append(cell_data)

        if len(cells) > 0:
            sheet_data['cells'] = cells

        workbook_data['worksheets'][selected_worksheet_index] = sheet_data

    return workbook_data


def elaborate_global_defined_names_description(defined_names):

    with open(workbook_description_output_file, 'w') as file:

        for name, data in defined_names.items():

            defined_name_type = data['type']

            description = f'The workbook has a {defined_name_type} named \"{name}\". '

            defined_name_range = data['range']

            match defined_name_type:
                case 'constant':
                    coordinate = f"'{defined_name_range['worksheet_title']}'!${defined_name_range['start']['col']}${defined_name_range['start']['row']}"
                    description += f"Its value is based on cell \"{coordinate}\"."
                case 'list' | 'table':
                    coordinate = f"'{defined_name_range['worksheet_title']}'!${defined_name_range['start']['col']}${defined_name_range['start']['row']}:${defined_name_range['end']['col']}${defined_name_range['end']['row']}"
                    description += f"Its value is based on range {coordinate}."

            file.write(description + '\n')


def elaborate_cells_description(worksheets_data):

    file_content = []

    for _, worksheet_data in worksheets_data.items():
        worksheet_name = worksheet_data['name']

        for cell in worksheet_data['cells']:

            coordinate = f"'{worksheet_name}'!{cell['coordinate']}"

            if 'value' in cell:
                description = f"contains the value \"{cell['value']}\""
                cell_type = 'value'

            if 'formula' in cell:
                description = f"contains the formula \"{cell['formula']}\""
                cell_type = 'formula'

            if 'require_output' in cell:
                description = 'requires an input from the user'
                cell_type = 'input'

            text = f'Cell {coordinate} {description}.'

            metadata = {
                'worksheet': worksheet_name,
                'coordinate': cell['coordinate'],
                'type': cell_type
            }

            if 'formula' in cell \
                    and 'inputs' in cell['formula']:
                inputs = cell['formula']['inputs']
                metadata['dependencies'] = []

                if 'cells' in inputs:
                    metadata['dependencies'].extend(inputs['cells'])

                if 'ranges' in inputs:
                    metadata['dependencies'].extend(inputs['ranges'])

                if 'defined_names' in inputs:
                    metadata['dependencies'].extend(
                        inputs['defined_names'])

            text = text.replace('\n', '\\n')\
                .replace('\t', '\\t')

            line = {
                'text': text,
                'metadata': metadata
            }

            file_content.append(line)

        with open(workbook_description_output_file, 'a') as file:
            json.dump(file_content, file, ensure_ascii=False,
                      indent=2, separators=(',', ': '))


def elaborate_workbook_description(workbook_data):

    elaborate_global_defined_names_description(workbook_data['defined_names'])

    elaborate_cells_description(workbook_data['worksheets'])


workbook_data = elaborate_workbook_data(spreadsheet_location)

output_file = os.path.join(output_directory, 'workbook-data.json')

with open(output_file, 'w') as file:
    json.dump(workbook_data, file, ensure_ascii=False,
              indent=2, separators=(',', ': '))


def elaborate_workbook_data():
    elaborate_workbook_description(workbook_data)
