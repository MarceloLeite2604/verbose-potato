import os
from configuration import load_env
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles.colors import COLOR_INDEX, Color
from openpyxl.styles.fills import PatternFill

load_env()


def retrieve_color(color: Color):
    if color.type == 'rgb':
        if color.rgb == '00000000':
            return None
        return color.rgb
    if color.type == 'indexed' \
            and color.index < 63:
        return str(COLOR_INDEX[color.index])
    return None


def elaborate_workbook_sections():
    spreadsheet_location = os.environ['INPUT_FILE_PATH']

    workbook = load_workbook(spreadsheet_location, data_only=False)

    for worksheet_index, worksheet in enumerate(workbook.worksheets):

        if worksheet_index != 3:
            continue

        # Pay attention to merged cells. They are described on the worksheet.
        merged_cells = worksheet.merged_cells

        print(f'Processing sheet {worksheet.title}')

        for row_index, row in enumerate(worksheet.iter_rows()):
            print(f'Processing row {row_index}.')

            for cell in row:

                if isinstance(cell, MergedCell):
                    continue

                font = {
                    'bold': cell.font.b,
                    'italic': cell.font.i,
                    'underline': cell.font.u if cell.font.u else False,
                    'font_size': cell.font.sz
                }

                borders = {
                    'top': True if cell.border.top.style else False,
                    'bottom': True if cell.border.bottom.style else False,
                    'left': True if cell.border.left.style else False,
                    'right': True if cell.border.right.style else False,
                }

                colors = {}

                if cell.font.color:
                    colors['font'] = retrieve_color(cell.font.color)

                if cell.fill.fgColor:
                    fg_color = retrieve_color(cell.fill.fgColor)
                    if fg_color:
                        colors['foreground'] = fg_color

                if cell.fill.bgColor:
                    bg_color = retrieve_color(cell.fill.bgColor)
                    if bg_color:
                        colors['background'] = bg_color

                print(colors)
                # if cell.value:
                #     bold = cell.font.bold
                #     italic = cell.font.italic
                #     font_size = cell.font.sz
