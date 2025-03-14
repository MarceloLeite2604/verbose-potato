import os
from dotenv import load_dotenv, find_dotenv
from xlsx2html import xlsx2html
import unidecode
from openpyxl import load_workbook

load_dotenv(find_dotenv(), override=True)

spreadsheet_location = os.environ['INPUT_FILE_PATH']

output_directory = 'output-files/html-content/'

# Create directory structure if it doesn't exist
if not os.path.exists(output_directory):
    os.makedirs(output_directory)

custom_spreadsheet_path = output_directory + 'custom-content.xlsx'

workbook = load_workbook(spreadsheet_location, data_only=True)

for index, worksheet in enumerate(workbook.worksheets):

    for row_index, _ in enumerate(worksheet.iter_rows()):
        row_dimentsions = worksheet.row_dimensions[row_index]
        if row_dimentsions.outline_level > 0:
            row_dimentsions.hidden = False

print(f'Saving workbook at \"{custom_spreadsheet_path}\".')
workbook.save(custom_spreadsheet_path)

for worksheet_index, worksheet in enumerate(workbook.worksheets):

    print(f'Generating HTML file for \"{worksheet.title}\" worksheet.')

    file_name = unidecode.unidecode(worksheet.title.replace(' ', '-').lower())

    file_path = output_directory + \
        str(worksheet_index) + '-' + file_name + '.html'

    xlsx2html(custom_spreadsheet_path, file_path,
              locale='pt_BR', sheet=worksheet_index)
